"""Build a Supabase seed SQL from Resort Info.xlsx.

Reads each resort sheet, detects section headers (GATE, CONTACTS, COMMUNITY
AMENITIES, POOL, TRASH, PARKING, PACKAGES, PETS, ELECTRIC CAR, etc.), pulls
their items into the JSONB sections shape the app expects, and writes a SQL
file that deletes-then-inserts every resort in a single transaction.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

HERE = Path(__file__).resolve().parent
REPO = HERE.parent
XLSX = HERE / "data" / "Resort Info.xlsx"
OUT = REPO / "db" / "seed_all_resorts.sql"

SKIP_SHEETS = {"Gates", "Sheet116", "Sheet127", "Sheet129", "Sheet130"}

# Canonical section mapping — matches the `type` values used in the existing
# Champions Gate seed + the dashboard renderer.
SECTION_MAP = {
    "GATE": ("gate", "Gate Access"),
    "CONTACTS": ("contacts", "Contacts"),
    "COMMUNITY AMENITIES": ("amenities", "Community Amenities"),
    "COMMUNITY AMENETIES": ("amenities", "Community Amenities"),
    "POOL": ("pool", "Pool"),
    "SPA": ("pool", "Spa"),
    "SPAS": ("pool", "Spa"),
    "TRASH": ("trash", "Trash"),
    "PARKING": ("parking", "Parking"),
    "• PARKING": ("parking", "Parking"),
    "PACKAGES": ("packages", "Packages"),
    "PETS": ("pets", "Pets"),
    "ELETRIC CAR": ("ev", "Electric Car"),
    "ELECTRIC CAR": ("ev", "Electric Car"),
    "CHARGER": ("ev", "Electric Car"),
    "ADDITIONAL INFO": ("additional", "Additional Info"),
}

PLACEHOLDERS = {"", "-", "--", "—", "?", "??", "???", "n/a", "na"}


def clean(v):
    if v is None:
        return ""
    s = str(v).replace("\r", " ").replace("\n", " | ").strip()
    s = re.sub(r"\s+\|\s+\|\s+", " | ", s)
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip()


def is_placeholder(s: str) -> bool:
    return s.strip().lower() in PLACEHOLDERS


def row_cells(row):
    """Return list of (col_letter, cleaned_value) for non-empty cells."""
    out = []
    for c in row:
        v = clean(c.value)
        if v:
            out.append((c.column_letter, v))
    return out


def is_header(text: str) -> bool:
    s = text.strip()
    if not s or len(s) > 40:
        return False
    if not any(ch.isalpha() for ch in s):
        return False
    # Uppercase only (treating non-letters as neutral)
    return s == s.upper() and not any(ch.islower() for ch in s)


def parse_sheet(ws):
    """Return dict: {canonical_type: {"title": ..., "items": [(label,value)...]}}."""
    sections: dict[str, dict] = {}
    current_type = None
    current_title = None
    col_headers: list[str] | None = None  # the "Location / Hours / Access / Court" row
    pending_after_header = False

    max_r = min(ws.max_row, 200)
    for row in ws.iter_rows(min_row=1, max_row=max_r):
        cells = row_cells(row)
        if not cells:
            col_headers = None
            continue

        # Detect a section header: single uppercase cell alone in its row.
        if len(cells) == 1 and is_header(cells[0][1]):
            key = cells[0][1].strip()
            if key in SECTION_MAP:
                current_type, current_title = SECTION_MAP[key]
                sections.setdefault(current_type, {"title": current_title, "items": []})
                col_headers = None
                pending_after_header = True
                continue
            # Unknown uppercase — treat as custom subsection under current
            if current_type is not None:
                sections[current_type]["items"].append(("", key))
            continue

        if current_type is None:
            # Likely the resort title at the top — ignore.
            continue

        values = [v for _, v in cells]
        # First row after header is usually column titles like "Login / Phone number / ..."
        if pending_after_header and len(values) >= 2 and all(
            len(v) < 40 and not is_placeholder(v) for v in values
        ):
            col_headers = values
            pending_after_header = False
            continue
        pending_after_header = False

        # Build an item
        useful = [v for v in values if not is_placeholder(v)]
        if not useful:
            continue

        if len(useful) == 1:
            label = ""
            value = useful[0]
        else:
            # If we have column headers and same arity, pair them up.
            if col_headers and len(useful) == len(col_headers):
                parts = [f"{h}: {v}" for h, v in zip(col_headers, useful)]
                label = useful[0]
                value = " — ".join(parts[1:]) if len(parts) > 1 else parts[0]
            else:
                label = useful[0]
                value = " — ".join(useful[1:])

        sections[current_type]["items"].append((label, value))

    # Drop empty sections
    sections = {k: v for k, v in sections.items() if v["items"]}
    return sections


GATE_CODE_RE = re.compile(r"\b(\d{3,6}#)")
GATE_CODE_LOOSE_RE = re.compile(r"(?:code|gate)[^\d]{0,20}(\d{3,6})#?", re.IGNORECASE)
ADDR_RE = re.compile(r"\b\d+[^,]{2,60},[^,]{2,40}FL[ ,]\d{5}", re.IGNORECASE)


def extract_gate_code(sections: dict) -> str | None:
    g = sections.get("gate")
    if not g:
        return None
    # Prefer explicit "####" with the hash marker.
    for label, value in g["items"]:
        m = GATE_CODE_RE.search(f"{label} {value}")
        if m:
            return m.group(1)
    # Otherwise require the word "code"/"gate" near the digits.
    for label, value in g["items"]:
        m = GATE_CODE_LOOSE_RE.search(f"{label} {value}")
        if m:
            return m.group(1)
    return None


def extract_address(sections: dict) -> str | None:
    c = sections.get("contacts")
    if not c:
        return None
    for label, value in c["items"]:
        for text in (label, value):
            m = ADDR_RE.search(text)
            if m:
                return m.group(0)
    # Fall back: first non-placeholder label that looks like an address
    for label, value in c["items"]:
        if re.search(r"\d+\s+\w+.*(Blvd|Dr|St|Ave|Rd|Way|Parkway|Pkwy|Circle|Cir|Ln|Ct)\b", label, re.IGNORECASE):
            return label
    return None


def clean_resort_name(sheet_name: str) -> tuple[str, bool]:
    """Return (display_name, is_out)."""
    s = sheet_name.strip()
    is_out = s.startswith("OUT - ")
    if is_out:
        s = s[len("OUT - "):].strip()
    return s, is_out


def make_aliases(name: str) -> list[str]:
    aliases = {name.lower()}
    # Common short tokens
    tokens = re.findall(r"[A-Za-z]+", name.lower())
    if len(tokens) >= 2:
        aliases.add("".join(t[0] for t in tokens))  # initials e.g. "cg"
    # Strip "the"
    if tokens and tokens[0] == "the":
        aliases.add(" ".join(tokens[1:]))
    return sorted(a for a in aliases if a)


def sections_to_json_list(sections: dict) -> list[dict]:
    """Convert to the app's JSONB shape: [{type,title,items:[{label,value}]}]."""
    # Preserve a stable order matching the existing seed.
    order = ["gate", "contacts", "amenities", "pool", "trash", "parking", "packages", "pets", "ev", "additional"]
    out = []
    for t in order:
        if t in sections:
            items = [{"label": l, "value": v} for l, v in sections[t]["items"]]
            out.append({"type": t, "title": sections[t]["title"], "items": items})
    # Any leftover types
    for t, data in sections.items():
        if t not in order:
            items = [{"label": l, "value": v} for l, v in data["items"]]
            out.append({"type": t, "title": data["title"], "items": items})
    return out


def sql_escape_literal(s: str) -> str:
    return s.replace("'", "''")


def sql_text_array(items: list[str]) -> str:
    parts = ", ".join(f"'{sql_escape_literal(i)}'" for i in items)
    return f"array[{parts}]::text[]" if parts else "array[]::text[]"


def build_insert(name: str, aliases: list[str], address: str | None, gate_code: str | None, sections_json: list[dict]) -> str:
    jsontxt = json.dumps(sections_json, ensure_ascii=False, indent=4)
    addr = f"'{sql_escape_literal(address)}'" if address else "null"
    gcode = f"'{sql_escape_literal(gate_code)}'" if gate_code else "null"
    return (
        f"    delete from public.resorts where name = '{sql_escape_literal(name)}';\n"
        f"    insert into public.resorts (name, aliases, address, gate_code, sections, updated_by) values (\n"
        f"        '{sql_escape_literal(name)}',\n"
        f"        {sql_text_array(aliases)},\n"
        f"        {addr},\n"
        f"        {gcode},\n"
        f"        $json${jsontxt}$json$::jsonb,\n"
        f"        'seed-xlsx'\n"
        f"    );\n"
    )


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    blocks = []
    stats = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        if ws.max_row < 5:
            continue
        name, is_out = clean_resort_name(sheet_name)
        sections = parse_sheet(ws)
        if not sections:
            stats.append((name, "skipped (no sections)"))
            continue
        aliases = make_aliases(name)
        if is_out:
            aliases.append("out of area")
        address = extract_address(sections)
        gate_code = extract_gate_code(sections)
        sections_json = sections_to_json_list(sections)
        blocks.append(build_insert(name, aliases, address, gate_code, sections_json))
        stats.append((name, f"{len(sections_json)} sections, gate={gate_code}, addr={'Y' if address else '-'}"))

    header = (
        "-- ============================================================\n"
        "-- Seed: All resorts (generated from Resort Info.xlsx)\n"
        "-- Run this in: Supabase Studio > SQL Editor\n"
        "-- Safe to re-run: each resort is delete-then-insert by name.\n"
        "-- ============================================================\n\n"
        "do $$\nbegin\n"
    )
    footer = "end $$;\n"
    OUT.write_text(header + "\n".join(blocks) + footer, encoding="utf-8")

    for name, info in stats:
        print(f"  {name:40s}  {info}")
    print(f"\nWrote {len(blocks)} resorts to {OUT}")


if __name__ == "__main__":
    main()
